#!/usr/bin/env python3
"""
Discontinued Items Finder - DTW to OITM Deactivation Tool
==========================================================

This script identifies items marked as DISCONTINUED in DTW files and creates
a deactivation file for DTW bulk update.

Process:
1. Reads DTW file (vendor price list)
2. Reads OITM file (current items in system)
3. Finds items that exist in BOTH files
4. Filters for items with "DISCONTINUED" in the Style Name (DTW file)
5. Creates DTW deactivation file with: ItemCode | frozenFor | validFor

Usage:
    python find_discontinued_items.py

File Requirements:
    - DTW files: Must contain 'DTW' or 'VPL' in filename
    - OITM files: Must contain 'OITM' in filename
    - DTW must have: Vendor Style, Style Name, Color, Size, Variable
    - OITM must have: Item No.

Output:
    - Deactivation DTW file with discontinued items ready for bulk upload
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from datetime import datetime


def parse_item_no(item_no):
    """Parse Item No. into Style, Color, Size, Variable components."""
    parts = str(item_no).split('-')
    
    if len(parts) == 4:
        return parts[0], parts[1], parts[2], parts[3]
    elif len(parts) == 3:
        return parts[0], parts[1], parts[2], None
    elif len(parts) > 4:
        style = parts[0]
        variable = parts[-1]
        size = parts[-2]
        color = '-'.join(parts[1:-2])
        return style, color, size, variable
    
    return None, None, None, None


def create_lookup_key(style, color, size, variable):
    """Create normalized lookup key for matching."""
    import pandas as pd
    
    style_n = str(style).upper().strip()
    color_n = str(color).upper().strip()
    size_n = str(size).upper().strip()
    
    # Handle Variable more carefully - check for empty, None, NaN, 'NONE', 'nan' string
    if pd.isna(variable) or variable is None or str(variable).strip() == '' or str(variable).upper().strip() in ['NONE', 'NAN']:
        var_n = ''
    else:
        var_n = str(variable).upper().strip()
    
    if var_n:
        return f"{style_n}|{color_n}|{size_n}|{var_n}"
    else:
        return f"{style_n}|{color_n}|{size_n}"


def normalize_dtw_columns(dtw_df):
    """Normalize DTW column names to expected format (case-insensitive)."""
    # Create mapping of lowercase column names to actual column names
    col_mapping = {str(col).lower().strip(): col for col in dtw_df.columns}
    
    # Required columns and their standardized names
    column_mappings = {
        'vendor style': 'Vendor Style',
        'color': 'Color',
        'size': 'Size',
        'variable': 'Variable',
        'style name': 'Style Name',
        'stylename': 'Style Name'
    }
    
    # Rename columns to standardized format
    rename_dict = {}
    for search_name, standard_name in column_mappings.items():
        if search_name in col_mapping:
            actual_col = col_mapping[search_name]
            if actual_col != standard_name:
                rename_dict[actual_col] = standard_name
    
    if rename_dict:
        dtw_df.rename(columns=rename_dict, inplace=True)
        print(f"  Normalized column names: {list(rename_dict.keys())} -> {list(rename_dict.values())}")
    
    return dtw_df


def find_discontinued_items(oitm_file, dtw_file, output_dir):
    """Find and create deactivation file for discontinued items."""
    
    vendor_name = oitm_file.stem.split('_')[0]
    
    print(f"\n{'='*80}")
    print(f"Processing {vendor_name}")
    print(f"{'='*80}")
    print(f"OITM File: {oitm_file.name}")
    print(f"DTW File:  {dtw_file.name}")
    
    try:
        # Read files
        oitm_df = pd.read_excel(oitm_file, header=1)  # Skip first header row
        dtw_df = pd.read_excel(dtw_file, header=0)  # Read from first row
        
        print(f"\nOriginal DTW columns: {list(dtw_df.columns)}")
        
        # Normalize DTW column names (case-insensitive)
        dtw_df = normalize_dtw_columns(dtw_df)
        
        print(f"OITM columns found: {list(oitm_df.columns)}")
        print(f"Normalized DTW columns: {list(dtw_df.columns)}")
        
        # Validate OITM - look for ItemCode column
        if 'ItemCode' not in oitm_df.columns:
            print(f"❌ ERROR: 'ItemCode' column not found in OITM file")
            print(f"   Available columns: {list(oitm_df.columns)}")
            return None
        
        # Validate DTW - check for Style Name column
        if 'Style Name' not in dtw_df.columns:
            print(f"❌ ERROR: 'Style Name' column not found in DTW file")
            print(f"   Available columns: {list(dtw_df.columns)}")
            return None
        
        # Required DTW columns
        required_cols = ['Vendor Style', 'Color', 'Size']
        missing_cols = [col for col in required_cols if col not in dtw_df.columns]
        if missing_cols:
            print(f"❌ ERROR: Missing columns in DTW file: {', '.join(missing_cols)}")
            print(f"   Available columns: {list(dtw_df.columns)}")
            return None
        
        # Parse OITM ItemCode
        print("\nParsing OITM items...")
        oitm_df[['Style', 'Color', 'Size', 'Variable']] = oitm_df['ItemCode'].apply(
            lambda x: pd.Series(parse_item_no(x))
        )
        
        # Create OITM lookup keys
        oitm_df['Lookup_Key'] = oitm_df.apply(
            lambda row: create_lookup_key(row['Style'], row['Color'], row['Size'], row['Variable']),
            axis=1
        )
        
        # Handle Variable column in DTW (might not exist)
        if 'Variable' not in dtw_df.columns:
            dtw_df['Variable'] = None
        
        # Create DTW lookup keys
        dtw_df['Lookup_Key'] = dtw_df.apply(
            lambda row: create_lookup_key(
                row['Vendor Style'], 
                row['Color'], 
                row['Size'], 
                row['Variable']
            ),
            axis=1
        )
        
        # Filter DTW for DISCONTINUED items (case-insensitive)
        print(f"\nSearching for DISCONTINUED items in 'Style Name' column...")
        discontinued_mask = dtw_df['Style Name'].astype(str).str.upper().str.contains('DISCONTINUED', na=False)
        discontinued_dtw = dtw_df[discontinued_mask].copy()
        
        print(f"Found {len(discontinued_dtw)} DISCONTINUED items in DTW file")
        
        if len(discontinued_dtw) == 0:
            print("✓ No discontinued items found - no deactivation needed")
            return {
                'vendor': vendor_name,
                'total_oitm': len(oitm_df),
                'discontinued_in_dtw': 0,
                'matched_to_deactivate': 0,
                'output_file': None
            }
        
        # Find matches between OITM and discontinued DTW items
        discontinued_lookup_keys = set(discontinued_dtw['Lookup_Key'])
        oitm_to_deactivate = oitm_df[oitm_df['Lookup_Key'].isin(discontinued_lookup_keys)].copy()
        
        # DEBUG: Show what we're looking for
        print(f"\n--- DEBUG INFO ---")
        print(f"Sample DISCONTINUED DTW items:")
        for idx, row in discontinued_dtw.head(3).iterrows():
            print(f"  Style:{row['Vendor Style']} Color:{row['Color']} Size:{row['Size']} Variable:{repr(row['Variable'])} -> Key:{row['Lookup_Key']}")
        
        print(f"\nSample OITM items:")
        for idx, row in oitm_df.head(10).iterrows():
            print(f"  ItemCode:{row['ItemCode']} -> Style:{row['Style']} Color:{row['Color']} Size:{row['Size']} Var:{repr(row['Variable'])} -> Key:{row['Lookup_Key']}")
        
        # Check if any discontinued styles exist in OITM
        if len(discontinued_dtw) > 0:
            sample_style = discontinued_dtw.iloc[0]['Vendor Style']
            print(f"\nDoes OITM contain style '{sample_style}'?")
            style_in_oitm = oitm_df[oitm_df['Style'] == str(sample_style)]
            print(f"  Found {len(style_in_oitm)} items with style {sample_style} in OITM")
            if len(style_in_oitm) > 0:
                print(f"  Sample ItemCodes:")
                for idx, row in style_in_oitm.head(5).iterrows():
                    print(f"    {row['ItemCode']} -> Key: {row['Lookup_Key']}")
        print(f"--- END DEBUG ---\n")
        
        print(f"Matched {len(oitm_to_deactivate)} items in OITM to deactivate")
        
        if len(oitm_to_deactivate) == 0:
            print("✓ No matching items in OITM - no deactivation needed")
            return {
                'vendor': vendor_name,
                'total_oitm': len(oitm_df),
                'discontinued_in_dtw': len(discontinued_dtw),
                'matched_to_deactivate': 0,
                'output_file': None
            }
        
        # Show sample of items to be deactivated
        print("\nSample items to be deactivated:")
        for idx, row in oitm_to_deactivate.head(5).iterrows():
            print(f"  - {row['ItemCode']}")
        if len(oitm_to_deactivate) > 5:
            print(f"  ... and {len(oitm_to_deactivate) - 5} more")
        
        # Create deactivation file
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Deactivate"
        
        # Dual headers for DTW
        headers = ['ItemCode', 'frozenFor', 'validFor']
        sheet.append(headers)
        sheet.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        for row_num in [1, 2]:
            for col_num in [1, 2, 3]:
                cell = sheet.cell(row=row_num, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add items to deactivate
        for idx, row in oitm_to_deactivate.iterrows():
            sheet.append([row['ItemCode'], 'Y', 'N'])
        
        # Format cells
        for row_num in range(3, sheet.max_row + 1):
            for col_num in [1, 2, 3]:
                cell = sheet.cell(row=row_num, column=col_num)
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths
        sheet.column_dimensions['A'].width = 35
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 15
        
        # Freeze headers
        sheet.freeze_panes = 'A3'
        
        # Save file
        output_path = output_dir / f"{vendor_name}_DEACTIVATE_DTW.xlsx"
        wb.save(output_path)
        
        print(f"\n✓ Deactivation file created: {output_path.name}")
        print(f"  Ready for DTW bulk upload in UPDATE mode")
        
        return {
            'vendor': vendor_name,
            'total_oitm': len(oitm_df),
            'discontinued_in_dtw': len(discontinued_dtw),
            'matched_to_deactivate': len(oitm_to_deactivate),
            'output_file': output_path.name,
            'items_list': oitm_to_deactivate['ItemCode'].tolist()
        }
        
    except Exception as e:
        print(f"❌ ERROR processing {vendor_name}: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


def find_file_pairs(input_dir):
    """Find matching OITM and DTW/VPL file pairs."""
    input_path = Path(input_dir)
    
    oitm_files = list(input_path.glob("*OITM*.xlsx"))
    dtw_files = list(input_path.glob("*DTW*.xlsx")) + list(input_path.glob("*VPL*.xlsx"))
    
    pairs = []
    
    for oitm_file in oitm_files:
        vendor_code = oitm_file.stem.split('_')[0]
        
        matching_dtw = None
        for dtw_file in dtw_files:
            if vendor_code.lower() in dtw_file.name.lower():
                matching_dtw = dtw_file
                break
        
        if matching_dtw:
            pairs.append((oitm_file, matching_dtw))
        else:
            print(f"⚠️  Warning: No DTW file found for {oitm_file.name}")
    
    return pairs


def create_summary_report(results, output_dir):
    """Create summary report of all deactivations."""
    
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Deactivation Summary"
    
    # Headers
    headers = ['Vendor', 'Total OITM Items', 'Discontinued in DTW', 'Items to Deactivate', 'Output File']
    sheet.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data
    total_deactivated = 0
    
    for result in results:
        if result:
            sheet.append([
                result['vendor'],
                result['total_oitm'],
                result['discontinued_in_dtw'],
                result['matched_to_deactivate'],
                result['output_file'] if result['output_file'] else 'No items to deactivate'
            ])
            total_deactivated += result['matched_to_deactivate']
    
    # Add totals row
    sheet.append([])
    sheet.append([
        'TOTAL',
        sum(r['total_oitm'] for r in results if r),
        sum(r['discontinued_in_dtw'] for r in results if r),
        total_deactivated,
        ''
    ])
    
    # Format cells
    for row_num in range(2, sheet.max_row + 1):
        for col_num in range(1, 6):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal='center' if col_num != 5 else 'left', vertical='center')
            
            if row_num == sheet.max_row:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    
    # Set column widths
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 25
    sheet.column_dimensions['D'].width = 25
    sheet.column_dimensions['E'].width = 40
    
    sheet.freeze_panes = 'A2'
    
    # Save report
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = output_dir / f"Deactivation_Summary_{timestamp}.xlsx"
    wb.save(report_path)
    
    return report_path


def main():
    """Main function to process all vendor files."""
    
    print("\n" + "="*80)
    print("DISCONTINUED ITEMS FINDER - DTW Deactivation Tool")
    print("="*80)
    
    # Set up directories
    input_dir = Path(r"C:\Users\it2\Downloads")
    output_dir = Path(r"C:\Users\it2\Downloads\results")
    
    output_dir.mkdir(exist_ok=True)
    
    print(f"\nInput Directory:  {input_dir}")
    print(f"Output Directory: {output_dir}")
    
    # Find file pairs
    print(f"\nSearching for OITM and DTW file pairs...")
    file_pairs = find_file_pairs(input_dir)
    
    if not file_pairs:
        print("\n❌ No matching OITM and DTW file pairs found!")
        return
    
    print(f"✓ Found {len(file_pairs)} vendor file pair(s)")
    
    # Process each vendor
    results = []
    for oitm_file, dtw_file in file_pairs:
        result = find_discontinued_items(oitm_file, dtw_file, output_dir)
        if result:
            results.append(result)
    
    # Create summary report
    if results:
        print(f"\n{'='*80}")
        print("Creating Summary Report")
        print(f"{'='*80}")
        
        report_path = create_summary_report(results, output_dir)
        print(f"✓ Summary report created: {report_path.name}")
        
        # Print final summary
        print(f"\n{'='*80}")
        print("PROCESSING COMPLETE")
        print(f"{'='*80}")
        print(f"Vendors processed: {len(results)}")
        print(f"Total items to deactivate: {sum(r['matched_to_deactivate'] for r in results):,}")
        print(f"\nAll output files saved to: {output_dir}")
        print(f"\nNext steps:")
        print(f"1. Review the deactivation files (*_DEACTIVATE_DTW.xlsx)")
        print(f"2. Import via DTW using OITM template in UPDATE mode")
        print(f"3. Map columns: ItemCode, frozenFor, validFor")
    else:
        print("\n❌ No vendors were successfully processed")


if __name__ == "__main__":
    main()