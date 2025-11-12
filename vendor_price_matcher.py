#!/usr/bin/env python3
"""
Vendor Price Matcher - Automated OITM Price Matching Tool
==========================================================

This script automatically matches prices from vendor VPL files to OITM files.

Usage:
    python vendor_price_matcher.py

File Requirements:
    - OITM files: Must have 'OITM' in filename and contain 'Item No.' column
    - VPL files: Must have 'VPL' or 'DTW' in filename and contain:
        - Vendor Style, Color, Size, Variable, Price columns

Output:
    - Updated OITM files with two header rows: ItemCode | U_VendorCost
    - Summary report of all processed vendors
    - Detailed list of removed SKUs (items without matching prices)
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os
from pathlib import Path
from datetime import datetime


# ============================================================================
# SIZE MAPPING CONFIGURATION
# ============================================================================

SIZE_MAPPING = {
    'XS': 'XSM',
    'S': 'SM',
    'M': 'MD',
    'L': 'LG',
    'XL': 'XLG',
    '2XL': '2XLG',
    '3XL': '3XLG',
    '4XL': '4XLG',
    '5XL': '5XLG'
}

# Define which styles use "G" suffix sizes
# Format: 'style' or ('style', 'color') for specific style-color combinations
STYLES_USING_G_SIZES = [
    '2278',           # Fabian Group style
    '3483',           # Rothco style
    ('2795', 'SILVER')  # Vantage Apparel style-color specific
]


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def parse_item_no(item_no):
    """Parse Item No. into Style, Color, Size, Variable components."""
    parts = str(item_no).split('-')
    
    if len(parts) == 4:
        # 4-part format: STYLE-COLOR-SIZE-VARIABLE
        style = parts[0]
        color = parts[1]
        size = parts[2]
        variable = parts[3]
        return style, color, size, variable
    elif len(parts) == 3:
        # 3-part format: STYLE-COLOR-SIZE
        style = parts[0]
        color = parts[1]
        size = parts[2]
        variable = None
        return style, color, size, variable
    elif len(parts) > 4:
        # More than 4 parts - color has hyphens
        style = parts[0]
        variable = parts[-1]
        size = parts[-2]
        color = '-'.join(parts[1:-2])
        return style, color, size, variable
    
    return None, None, None, None


def apply_conditional_size_mapping(row, styles_using_g_sizes):
    """Apply size mapping only for specific styles or style-color combinations."""
    style = row['Style']
    color = row['Color']
    size = row['Size']
    
    # Check if style-color combo needs mapping
    for config in styles_using_g_sizes:
        if isinstance(config, tuple):
            # Style-Color specific mapping
            if (style, color) == config:
                return SIZE_MAPPING.get(size, size)
        else:
            # Style-only mapping
            if style == config:
                return SIZE_MAPPING.get(size, size)
    
    return size


def create_output_file(oitm_filtered, vendor_name, output_dir):
    """Create the formatted Excel output file."""
    wb = Workbook()
    sheet = wb.active
    sheet.title = f"{vendor_name}_Updated"
    
    # TWO HEADER ROWS
    header_row_1 = ['ItemCode', 'U_VendorCost']
    header_row_2 = ['ItemCode', 'U_VendorCost']
    
    sheet.append(header_row_1)
    sheet.append(header_row_2)
    
    # Style both header rows
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    
    for row_num in [1, 2]:
        for col_num in [1, 2]:
            cell = sheet.cell(row=row_num, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data rows (only rows with prices)
    for idx, row in oitm_filtered.iterrows():
        sheet.append([row['Item No.'], row['Price']])
    
    # Format the data
    for row_num in range(3, sheet.max_row + 1):
        # ItemCode cell
        item_cell = sheet.cell(row=row_num, column=1)
        item_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # U_VendorCost cell
        price_cell = sheet.cell(row=row_num, column=2)
        price_cell.number_format = '0.00'
        price_cell.alignment = Alignment(horizontal='right', vertical='center')
    
    # Set column widths
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 18
    
    # Freeze top 2 header rows
    sheet.freeze_panes = 'A3'
    
    # Save file
    output_path = output_dir / f"{vendor_name}_OITM_Updated.xlsx"
    wb.save(output_path)
    
    return output_path


def match_prices(oitm_df, vpl_df, styles_using_g_sizes):
    """Match prices from VPL to OITM with size mapping support."""
    
    # Parse OITM Item No.
    oitm_df[['Style', 'Color', 'Size', 'Variable']] = oitm_df['Item No.'].apply(
        lambda x: pd.Series(parse_item_no(x))
    )
    
    # Apply size mapping conditionally
    oitm_df['Size_Mapped'] = oitm_df.apply(
        lambda row: apply_conditional_size_mapping(row, styles_using_g_sizes),
        axis=1
    )
    
    # Normalize OITM fields
    oitm_df['Style_norm'] = oitm_df['Style'].astype(str).str.upper().str.strip()
    oitm_df['Color_norm'] = oitm_df['Color'].astype(str).str.upper().str.strip()
    oitm_df['Size_norm'] = oitm_df['Size'].astype(str).str.upper().str.strip()
    oitm_df['Size_Mapped_norm'] = oitm_df['Size_Mapped'].astype(str).str.upper().str.strip()
    oitm_df['Variable_norm'] = oitm_df['Variable'].fillna('').astype(str).str.upper().str.strip()
    
    # Normalize VPL fields
    vpl_df['Style_norm'] = vpl_df['Vendor Style'].astype(str).str.upper().str.strip()
    vpl_df['Color_norm'] = vpl_df['Color'].astype(str).str.upper().str.strip()
    vpl_df['Size_norm'] = vpl_df['Size'].astype(str).str.upper().str.strip()
    vpl_df['Variable_norm'] = vpl_df['Variable'].fillna('').astype(str).str.upper().str.strip()
    
    # Create lookup keys
    vpl_df['Lookup_Key_4'] = (
        vpl_df['Style_norm'] + '|' + 
        vpl_df['Color_norm'] + '|' + 
        vpl_df['Size_norm'] + '|' +
        vpl_df['Variable_norm']
    )
    
    vpl_df['Lookup_Key_3'] = (
        vpl_df['Style_norm'] + '|' + 
        vpl_df['Color_norm'] + '|' + 
        vpl_df['Size_norm']
    )
    
    oitm_df['Lookup_Key_4_Mapped'] = (
        oitm_df['Style_norm'] + '|' + 
        oitm_df['Color_norm'] + '|' + 
        oitm_df['Size_Mapped_norm'] + '|' +
        oitm_df['Variable_norm']
    )
    
    oitm_df['Lookup_Key_3_Mapped'] = (
        oitm_df['Style_norm'] + '|' + 
        oitm_df['Color_norm'] + '|' + 
        oitm_df['Size_Mapped_norm']
    )
    
    # Create lookup dictionaries
    price_lookup_4 = dict(zip(vpl_df['Lookup_Key_4'], vpl_df['Price']))
    price_lookup_3 = dict(zip(vpl_df['Lookup_Key_3'], vpl_df['Price']))
    
    # Match prices with MAPPED sizes
    oitm_df['Price'] = oitm_df['Lookup_Key_4_Mapped'].map(price_lookup_4)
    no_match_mask = oitm_df['Price'].isna()
    oitm_df.loc[no_match_mask, 'Price'] = oitm_df.loc[no_match_mask, 'Lookup_Key_3_Mapped'].map(price_lookup_3)
    
    return oitm_df


def process_vendor(oitm_file, vpl_file, output_dir, styles_using_g_sizes):
    """Process a single vendor's OITM and VPL files."""
    
    vendor_name = oitm_file.stem.split('_')[0]  # Extract vendor code (e.g., V105 from V105_OITM.xlsx)
    
    print(f"\n{'='*70}")
    print(f"Processing {vendor_name}")
    print(f"{'='*70}")
    print(f"OITM File: {oitm_file.name}")
    print(f"VPL File:  {vpl_file.name}")
    
    try:
        # Read files
        oitm_df = pd.read_excel(oitm_file)
        vpl_df = pd.read_excel(vpl_file)
        
        # Validate columns
        if 'Item No.' not in oitm_df.columns:
            print(f"❌ ERROR: 'Item No.' column not found in OITM file")
            return None
        
        required_vpl_cols = ['Vendor Style', 'Color', 'Size', 'Variable', 'Price']
        missing_cols = [col for col in required_vpl_cols if col not in vpl_df.columns]
        if missing_cols:
            print(f"❌ ERROR: Missing columns in VPL file: {', '.join(missing_cols)}")
            return None
        
        # Match prices
        oitm_df = match_prices(oitm_df, vpl_df, styles_using_g_sizes)
        
        # Statistics
        total_skus = len(oitm_df)
        matched_skus = oitm_df['Price'].notna().sum()
        removed_skus = total_skus - matched_skus
        match_rate = (matched_skus / total_skus * 100) if total_skus > 0 else 0
        
        # Count size-mapped items
        size_mapped_count = len(oitm_df[(oitm_df['Price'].notna()) & (oitm_df['Size'] != oitm_df['Size_Mapped'])])
        
        # Get list of removed items
        removed_items = oitm_df[oitm_df['Price'].isna()]['Item No.'].tolist()
        
        print(f"\nResults:")
        print(f"  Total SKUs:          {total_skus:,}")
        print(f"  Matched SKUs:        {matched_skus:,} ({match_rate:.1f}%)")
        if size_mapped_count > 0:
            print(f"  Size-mapped SKUs:    {size_mapped_count:,}")
        print(f"  Removed (no price):  {removed_skus:,} ({removed_skus/total_skus*100:.1f}%)")
        
        # Filter to only matched rows
        oitm_filtered = oitm_df[oitm_df['Price'].notna()].copy()
        
        # Create output file
        output_path = create_output_file(oitm_filtered, vendor_name, output_dir)
        print(f"\n✓ Output file created: {output_path.name}")
        
        # Return summary data
        return {
            'vendor': vendor_name,
            'total_skus': total_skus,
            'matched_skus': matched_skus,
            'removed_skus': removed_skus,
            'match_rate': match_rate,
            'size_mapped': size_mapped_count,
            'output_file': output_path.name,
            'removed_items': removed_items  # NEW: List of removed Item Numbers
        }
        
    except Exception as e:
        print(f"❌ ERROR processing {vendor_name}: {str(e)}")
        return None


def find_file_pairs(input_dir):
    """Find matching OITM and VPL/DTW file pairs."""
    input_path = Path(input_dir)
    
    oitm_files = list(input_path.glob("*OITM*.xlsx"))
    vpl_files = list(input_path.glob("*VPL*.xlsx")) + list(input_path.glob("*DTW*.xlsx"))
    
    pairs = []
    
    for oitm_file in oitm_files:
        # Extract vendor code (e.g., V105, V106, etc.)
        vendor_code = oitm_file.stem.split('_')[0]
        
        # Find matching VPL file
        matching_vpl = None
        for vpl_file in vpl_files:
            if vendor_code.lower() in vpl_file.name.lower():
                matching_vpl = vpl_file
                break
        
        if matching_vpl:
            pairs.append((oitm_file, matching_vpl))
        else:
            print(f"⚠️  Warning: No VPL file found for {oitm_file.name}")
    
    return pairs


def create_summary_report(results, output_dir):
    """Create a summary report of all processed vendors with removed SKUs list."""
    
    wb = Workbook()
    
    # ========================================================================
    # SHEET 1: Processing Summary
    # ========================================================================
    summary_sheet = wb.active
    summary_sheet.title = "Processing Summary"
    
    # Headers
    headers = ['Vendor', 'Total SKUs', 'Matched SKUs', 'Match Rate %', 'Size Mapped', 'Removed', 'Output File']
    summary_sheet.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    for col_num, header in enumerate(headers, 1):
        cell = summary_sheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data
    total_skus_all = 0
    total_matched_all = 0
    
    for result in results:
        if result:
            summary_sheet.append([
                result['vendor'],
                result['total_skus'],
                result['matched_skus'],
                f"{result['match_rate']:.1f}%",
                result['size_mapped'] if result['size_mapped'] > 0 else '',
                result['removed_skus'],
                result['output_file']
            ])
            total_skus_all += result['total_skus']
            total_matched_all += result['matched_skus']
    
    # Add totals row
    summary_sheet.append([])
    summary_sheet.append([
        'TOTAL',
        total_skus_all,
        total_matched_all,
        f"{total_matched_all/total_skus_all*100:.1f}%" if total_skus_all > 0 else "0%",
        '',
        total_skus_all - total_matched_all,
        ''
    ])
    
    # Format cells
    for row_num in range(2, summary_sheet.max_row + 1):
        for col_num in range(1, 8):
            cell = summary_sheet.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal='center' if col_num != 7 else 'left', vertical='center')
            
            # Bold totals row
            if row_num == summary_sheet.max_row:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    
    # Set column widths
    summary_sheet.column_dimensions['A'].width = 15
    summary_sheet.column_dimensions['B'].width = 15
    summary_sheet.column_dimensions['C'].width = 15
    summary_sheet.column_dimensions['D'].width = 15
    summary_sheet.column_dimensions['E'].width = 15
    summary_sheet.column_dimensions['F'].width = 15
    summary_sheet.column_dimensions['G'].width = 35
    
    # ========================================================================
    # SHEET 2: Removed SKUs by Vendor
    # ========================================================================
    removed_sheet = wb.create_sheet("Removed SKUs")
    
    # Headers for removed SKUs sheet
    removed_headers = ['Vendor', 'Item No.', 'Reason']
    removed_sheet.append(removed_headers)
    
    # Style headers
    removed_header_fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    removed_header_font = Font(bold=True, color='FFFFFF', size=12)
    for col_num, header in enumerate(removed_headers, 1):
        cell = removed_sheet.cell(row=1, column=col_num)
        cell.font = removed_header_font
        cell.fill = removed_header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add removed items for each vendor
    current_row = 2
    for result in results:
        if result and result['removed_items']:
            vendor_name = result['vendor']
            
            # Add vendor section header
            vendor_cell = removed_sheet.cell(row=current_row, column=1)
            vendor_cell.value = vendor_name
            vendor_cell.font = Font(bold=True, size=11)
            vendor_cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
            vendor_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Merge cells for vendor header
            removed_sheet.merge_cells(start_row=current_row, start_column=1, 
                                     end_row=current_row, end_column=3)
            current_row += 1
            
            # Add each removed item
            for item_no in result['removed_items']:
                removed_sheet.append([vendor_name, item_no, 'No matching price found'])
                current_row += 1
            
            # Add blank row between vendors
            current_row += 1
    
    # Format removed SKUs sheet
    for row_num in range(2, removed_sheet.max_row + 1):
        for col_num in range(1, 4):
            cell = removed_sheet.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal='left' if col_num == 2 else 'center', vertical='center')
    
    # Set column widths for removed SKUs sheet
    removed_sheet.column_dimensions['A'].width = 15
    removed_sheet.column_dimensions['B'].width = 35
    removed_sheet.column_dimensions['C'].width = 30
    
    # Freeze top row in both sheets
    summary_sheet.freeze_panes = 'A2'
    removed_sheet.freeze_panes = 'A2'
    
    # Save report
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = output_dir / f"Processing_Summary_{timestamp}.xlsx"
    wb.save(report_path)
    
    return report_path


# ============================================================================
# MAIN FUNCTION
# ============================================================================

def main():
    """Main function to process all vendor files."""
    
    print("\n" + "="*70)
    print("VENDOR PRICE MATCHER - Automated Processing")
    print("="*70)
    
    # Set up directories
    input_dir = Path(r"C:\Users\it2\Downloads")
    output_dir = Path(r"C:\Users\it2\Downloads\results")
    
    output_dir.mkdir(exist_ok=True)
    
    print(f"\nInput Directory:  {input_dir}")
    print(f"Output Directory: {output_dir}")
    
    # Find file pairs
    print(f"\nSearching for OITM and VPL/DTW file pairs...")
    file_pairs = find_file_pairs(input_dir)
    
    if not file_pairs:
        print("\n❌ No matching OITM and VPL/DTW file pairs found!")
        print("\nExpected file naming:")
        print("  - OITM files: Must contain 'OITM' (e.g., V105_OITM.xlsx)")
        print("  - VPL files: Must contain 'VPL' or 'DTW' (e.g., V105_VPL.xlsx, V105_DTW.xlsx)")
        return
    
    print(f"✓ Found {len(file_pairs)} vendor file pair(s)")
    
    # Process each vendor
    results = []
    for oitm_file, vpl_file in file_pairs:
        result = process_vendor(oitm_file, vpl_file, output_dir, STYLES_USING_G_SIZES)
        if result:
            results.append(result)
    
    # Create summary report
    if results:
        print(f"\n{'='*70}")
        print("Creating Summary Report")
        print(f"{'='*70}")
        
        report_path = create_summary_report(results, output_dir)
        print(f"✓ Summary report created: {report_path.name}")
        print(f"  - Sheet 1: Processing Summary")
        print(f"  - Sheet 2: Removed SKUs (detailed list by vendor)")
        
        # Print final summary
        print(f"\n{'='*70}")
        print("PROCESSING COMPLETE")
        print(f"{'='*70}")
        print(f"Vendors processed: {len(results)}")
        print(f"Total SKUs: {sum(r['total_skus'] for r in results):,}")
        print(f"Total matched: {sum(r['matched_skus'] for r in results):,}")
        print(f"Total removed: {sum(r['removed_skus'] for r in results):,}")
        print(f"Overall match rate: {sum(r['matched_skus'] for r in results)/sum(r['total_skus'] for r in results)*100:.1f}%")
        print(f"\nAll output files saved to: {output_dir}")
    else:
        print("\n❌ No vendors were successfully processed")


if __name__ == "__main__":
    main()